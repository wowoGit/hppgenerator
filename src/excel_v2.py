#!/usr/bin/env python3
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

           
class TableForm:
    def __init__(self) -> None:
        self._interface = None
        self._unitname = None
        self._prototype = None
        self._parameters = None
        self._returnval = None
        self._description = None
        self.header = 'Software Unit Information'
        self.table = {
                "Interface" : None,
                "Unit Name": None,
                "Prototype" : None,
                "Parameters": [
                    [
                        "Type", "Name", "Value/Range", "IN/OUT", "Description",
                    ],
                ],
                'Return Type': [
                    [
                    'Type', 'Value/Range'
                    ]
                ],
                'Description': None,
                'Global Variables': [
                    ['Name', 'Read/Write'],
                    ['-', '-']
                ],
                'Called Function': '-',
                'Calling Function': '-',
                'Execution Time': '-'

    
        }
    

    def _get_interface(self):
        return self.table.get('Interface')

    def _set_interface(self, value):
        self.table['Interface'] = value

    def _get_returntype(self):
        return self.table.get('Return Type')

    def _set_returntype(self, value):
        if len(self.table['Return Type']) > 1:
            self.table['Return Type'][1] = value
            return
        
        self.table['Return Type'].append(value)


    def _get_unitname(self):
        return self.table.get('Unit Name')

    def _set_unitname(self, value):
        self.table['Unit Name'] = value

    def _get_prototype(self):
        return self.table.get('Prototype')

    def _set_prototype(self, value):
        self.table['Prototype'] = value

    def _get_parameters(self):
        return self.table.get('Parameters')

    def _set_parameters(self, value):
        self.table['Parameters'].append(value)

    def _get_description(self):
        return self.table.get('Description')

    def _set_description(self, value):
        self.table['Description'] = value

    inteface = property(
        fget=_get_interface,
        fset=_set_interface,
        doc="The interface property."
    )
    description = property(
        fget=_get_description,
        fset=_set_description,
        doc="The description property."
    )
    returntype = property(
        fget=_get_returntype,
        fset=_set_returntype,
        doc="The return type property."
    )
    unitname = property(
        fget=_get_unitname,
        fset=_set_unitname,
        doc="The Unit Name property."
    )
    prototype = property(
        fget=_get_prototype,
        fset=_set_prototype,
        doc="The prototype property."
    )
    parameters = property(
        fget=_get_parameters,
        fset=_set_parameters,
        doc="The parameters property."
    )
class Table:
    def __init__(self,workbook: Workbook,classname,class_description ):
        self.workbook = workbook
        self.class_name = classname
        self.worksheet = self.workbook.create_sheet(classname)
        self.class_desc = class_description
        self.border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
        self.fill = PatternFill(start_color='FFB2B2B2',
                                end_color='FFB2B2B2',
                                fill_type='solid')
        self.createFormat()
        self.fillHeader()

    def fillHeader(self):
        self.worksheet.cell(row=1,column=1).value = 'Class'
        self.worksheet.cell(row=1,column=2).value = self.class_name
        self.worksheet.cell(row=2,column=1).value = 'Description'
        self.worksheet.cell(row=2,column=2).value = self.class_desc
        self.setBorder(1,2,1,2,self.border)
        self.setFill(1,2,1,1,self.fill)
        self.setAlignment(1,2,1,1, Alignment(horizontal='center',vertical='center',wrap_text=True))
        self.setAlignment(1,2,2,2, Alignment(wrap_text=True))

    def createFormat(self):
        self.worksheet.column_dimensions['A'].width = 25.67
        self.worksheet.column_dimensions['B'].width = 25.78
        self.worksheet.column_dimensions['C'].width = 20.89
        self.worksheet.column_dimensions['D'].width = 28.56
        self.worksheet.column_dimensions['E'].width = 9.22
        self.worksheet.column_dimensions['F'].width = 56.78

    def toExcel(self,form:TableForm):
        ws = self.worksheet
        table_spacing = 3
        row = ws.max_row + table_spacing
        column = 2
        table_start_row = row
        ws.cell(row,1).value = form.header
        ws.merge_cells(start_row=row, start_column=1, end_row=row,end_column=6)
        ws.cell(row=row,column=1).alignment = Alignment(horizontal='center')
        self.setAlignment(row,row,1,1, Alignment(horizontal='center',vertical='center',wrap_text=True))
        row += 1
        for k, v in form.table.items():
            ws.cell(row, 1).value = k
            ws.cell(row,1).alignment = Alignment(horizontal='center')
            self.setAlignment(row,row,1,1, Alignment(horizontal='center',vertical='center',wrap_text=True))
            if type(v) == list:
                ws.merge_cells(start_row=row, start_column=1, end_row=row + len(v) - 1,end_column=1)
                row_old = row
                for inner in v:
                    for val in inner:
                        if val in v[0]:
                            ws.cell(row, column).fill = self.fill
                            self.setAlignment(row,row,column,column, Alignment(horizontal='center',vertical='center',wrap_text=True))
                            self.setFill(row,row,1,6,self.fill)
                        else:
                            self.setAlignment(row,row,column,column,Alignment(wrap_text=True))
                        ws.cell(row, column).value = val
                        column += 1
                        if k == 'Return Type' or k == 'Global Variables':
                            column +=1
                    column = 2
                    row+=1
                if k == 'Return Type' or k == 'Global Variables':
                    merge_row = row_old
                    while merge_row < row:
                        ws.merge_cells(start_row=merge_row, start_column=2, end_row=merge_row,end_column = 3)
                        ws.merge_cells(start_row=merge_row, start_column=4, end_row=merge_row,end_column = 6)
                        merge_row += 1
            else:
                ws.cell(row, column).value = v
                self.setAlignment(row,row,column,column,Alignment(wrap_text=True))
                ws.merge_cells(start_row=row, start_column=2, end_row=row,end_column = 6)
                row +=1

        self.setBorder(row_start = table_start_row, row_end=row-1,col_start=1,col_end=6,border=self.border)
        self.setFill(table_start_row,row - 1,1,1,self.fill)
        row += table_spacing

    def setBorder(self, row_start, row_end, col_start, col_end,border : Border):
        ws = self.worksheet
        for row in ws.iter_rows(min_row=row_start, min_col=col_start, max_row=row_end, max_col=col_end):
            for cell in row:
                cell.border = border
    def setAlignment(self, row_start, row_end, col_start, col_end,alignment):
        ws = self.worksheet
        for row in ws.iter_rows(min_row=row_start, min_col=col_start, max_row=row_end, max_col=col_end):
            for cell in row:
                cell.alignment = alignment

    def setFill(self, row_start, row_end, col_start, col_end,fill : PatternFill):
        ws = self.worksheet
        for row in ws.iter_rows(min_row=row_start, min_col=col_start, max_row=row_end, max_col=col_end):
            for cell in row:
                cell.fill = fill

def main():
    wb = Workbook()
    wb.remove(wb.active)
    tbl = Table(wb, "<classname>",'<class description>')
    tb = TableForm()
    tb.inteface = 'YES'
    tb.unitname = 'blablabla'
    tb.prototype = 'void blablabla(int a)'
    tb.description = 'method description'
    tb.parameters.append(['int','a','0-255','IN','Param Description'])
    tbl.toExcel(tb)
    wb.save(os.path.join(os.getcwd(),'result.xlsx'))
if __name__ == '__main__':
    main()


