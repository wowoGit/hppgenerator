#!/usr/bin/env python3
import xml.etree.ElementTree as ET
import openpyxl
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
import argparse

#sections that should be exported to excel
export_to_excel = [
    'public-func',
    'public-slot',
    'signal',
    'event',
    'public-static-func',
    'protected-func',
    'protected-static-func',
    'protected-slot',
    'private-func',
    'private-static-func',
    'friend',
    'prototype'
]

def str_if_none(val):
    return val or ''

def unpack_ref(node: ET.Element):
    ref = node[0]
    return str_if_none(node.text) + ref.text + str_if_none(ref.tail)

def removescope(name :str):
    namespace_end = name.find('::')
    return name[namespace_end+2:] if namespace_end != -1 else name

class XmlObject:
    def __init__(self,node: ET.Element):
        self.xmlnode = node
        self.formBrief()


    def formBrief(self) :
        self.brief = None
        if self.containsField(self.xmlnode,'briefdescription'):
            brief_node = self.xmlnode.find('briefdescription')
            for field in list(brief_node):
                self.brief = ''
                if self.containsField(field,'ref'):
                    ref = list(field)[0]
                    self.brief += str(field.text or '') + ref.text + ref.tail
                else:
                    self.brief = field.text
    
    def containsField(self,node: ET.Element, fieldname:str):
        if node.find(fieldname) is not None:
            return True
        return False

class XmlFunc(XmlObject):
    def __init__(self, node: ET.Element):
        super().__init__(node)
        self.params_desc = {}
        self.returnval = {}
        self.name = ''
        self.definition = ''
        self.template = ''
        self.protection = node.get('prot')
        self.formTemplate()
        self.formDefinition()
        self.formParams()

    def formTemplate(self):
        self.template = None
        if self.containsField(self.xmlnode,'templateparamlist'):
            templatelist = self.xmlnode.find('templateparamlist')
            self.template = "<"
            for field in list(templatelist):
                self.template += field.findtext('type') + ' '
                if self.containsField(field,'declname'):
                    self.template += ' ' + field.findtext('declname')
            self.template += '>'

    def formParams(self):
        def getRange(string: str):
                range_keyword_start = string.find('Range')
                range_end = str(string[range_keyword_start:]).find('.') + range_keyword_start
                range_text = string[range_keyword_start + len('Range '):range_end]
                stripped_string = string[:range_keyword_start] + string[range_end+1:]
                return stripped_string,range_text

        if self.containsField(self.xmlnode,'detaileddescription'):
            detailed_node = self.xmlnode.find('detaileddescription')
            for para in list(detailed_node):
                for child in list(para):
                    if child.get('kind') == 'param':
                        for field in list(child):
                                param_name = field.find('parameternamelist/parametername')
                                param_name_text = param_name.text
                                param_dir = param_name.get('direction')
                                param_desc_text = field.find('parameterdescription/para').text
                                range_text = None
                                if param_desc_text.__contains__('Range'):
                                    param_desc_text,range_text = getRange(param_desc_text)
                                param = [param for param in self.xmlnode.findall('param') if param.findtext('declname') ==param_name_text][0]
                                param_type = param.find('type')
                                param_type_text = param_type.text or ''
                                if self.containsField(param_type,'ref'):
                                    param_type_text += unpack_ref(param_type)
                                self.params_desc[param_name_text] = { 'value/range':range_text,'type':param_type_text, "direction":param_dir, 'description': param_desc_text}
                    if child.get('kind') == 'return':
                        simplesect_text = str_if_none(child.find('para').text)
                        if self.containsField(child.find('para'),'ref'):
                            simplesect_text += unpack_ref(child.find('para'))
                        range_text = None
                        if simplesect_text.__contains__('Range'):
                            simplesect_text,range_text = getRange(simplesect_text) # range will be stripped from string
                        if self.containsField(self.xmlnode.find('type'),'ref'):
                            self.returnval['type'] = unpack_ref(self.xmlnode.find('type'))
                        else:
                            self.returnval['type'] = self.xmlnode.findtext('type')
                        self.returnval['description'] = simplesect_text
                        self.returnval['range/value'] = range_text
                    if child.get('kind') == 'see':
                        para = child.find('para')
                        self.brief += para.findtext('ref')


    def formDefinition(self):
        func_def = ""
        full_def = self.xmlnode.find('definition').text
        func_name = self.xmlnode.find('name').text
        self.name = func_name
        if full_def.startswith(func_name) or func_name.startswith('~'):
            if self.xmlnode.get('explicit') == 'yes':
                func_def += 'explicit '
            func_def += func_name
        else:
            func_def += full_def 
        if self.containsField(self.xmlnode,'argsstring'):
            func_def += self.xmlnode.findtext('argsstring')
        if self.template is not None:
            arg_pos = func_def.find('(')
            func_def = func_def[:arg_pos] + self.template + func_def[arg_pos:]
            self.name +=self.template
        self.definition = func_def

class XmlSection(XmlObject):
    def __init__(self,xmlnode: ET.Element):
        super().__init__(xmlnode)
        self.kind = xmlnode.get('kind')
        self.fields = [XmlFunc(node) for node in self.xmlnode.findall('memberdef')]


class XmlClass(XmlObject):
    def __init__(self,xmlnode):
        doc = ET.parse(xmlnode)
        root = doc.getroot()
        super().__init__(root.find('compounddef'))
        self.name = removescope(self.xmlnode.findtext('compoundname'))
        self.sections = [XmlSection(node) for node in self.xmlnode.findall('sectiondef') if node.get('kind') in export_to_excel]
    


           
class TableForm:
    def __init__(self) -> None:
        self.header = 'Software Unit Information'
        self.table = {
                "Interface" : None,
                "Unit Name": None,
                "Prototype" : None,
                "Parameters": [
                    [
                        "Type", "Name", "Value/Range", "IN/OUT", "Description",
                    ],
                ],
                'Return Type': [
                    [
                    'Type', 'Value/Range'
                    ]
                ],
                'Description': None,
                'Global Variables': [
                    ['Name', 'Read/Write'],
                    ['-', '-']
                ],
                'Called Function': '-',
                'Calling Function': '-',
                'Execution Time': '-'

    
        }
    
    @property
    def interface(self):
        return self.table.get('Interface')

    @interface.setter
    def interface(self, value):
        self.table['Interface'] = value

    @property
    def returntype(self):
        return self.table.get('Return Type')

    @returntype.setter
    def returntype(self, value):
        if len(self.table['Return Type']) > 1:
            self.table['Return Type'][1] = value
            return
        
        self.table['Return Type'].append(value)

    @property
    def unitname(self):
        return self.table.get('Unit Name')

    @unitname.setter
    def unitname(self, value):
        self.table['Unit Name'] = value

    @property
    def prototype(self):
        return self.table.get('Prototype')

    @prototype.setter
    def prototype(self, value):
        self.table['Prototype'] = value

    @property
    def parameters(self):
        return self.table.get('Parameters')

    @parameters.setter
    def parameters(self, value):
        self.table['Parameters'].append(value)

    @property
    def description(self):
        return self.table.get('Description')

    @description.setter
    def description(self, value):
        self.table['Description'] = value

class Table:
    def __init__(self,workbook: Workbook, xmlclass: XmlClass):
            self.workbook = workbook
            self.xmlclass = xmlclass
            self.worksheet = self.workbook.create_sheet(self.xmlclass.name)
            self.border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            self.fill = PatternFill(start_color='FFB2B2B2',
                                    end_color='FFB2B2B2',
                                    fill_type='solid')
            self.applySize()

    def applySize(self):
        self.worksheet.column_dimensions['A'].width = 25.67
        self.worksheet.column_dimensions['B'].width = 25.78
        self.worksheet.column_dimensions['C'].width = 20.89
        self.worksheet.column_dimensions['D'].width = 28.56
        self.worksheet.column_dimensions['E'].width = 9.22
        self.worksheet.column_dimensions['F'].width = 56.78

    def fillHeader(self):
        self.worksheet.cell(row=1,column=1).value = 'Class'
        self.worksheet.cell(row=1,column=2).value = self.xmlclass.name
        self.worksheet.cell(row=2,column=1).value = 'Description'
        self.worksheet.cell(row=2,column=2).value = self.xmlclass.brief
        self.setBorder(1,2,1,2,self.border)
        self.setFill(1,2,1,1,self.fill)
        self.setAlignment(1,2,1,1, Alignment(horizontal='center',vertical='center',wrap_text=True))
        self.setAlignment(1,2,2,2, Alignment(wrap_text=True))

    def fillForm(self,field:XmlFunc) -> TableForm:
        form = TableForm()
        form.interface = 'YES' if field.protection == 'public' else 'NO'
        form.unitname = field.name
        form.prototype = field.definition
        form.description = field.brief
        if len(field.returnval) > 0:
            form.returntype = [field.returnval['type'], field.returnval['range/value']]
        else:
                form.returntype = ['-','-']
        if len(field.params_desc) > 0:
            for param_key_name,param_properties in field.params_desc.items():
                
                form.parameters.append([param_properties['type'], # should be strictly in this order
                                        param_key_name,
                                        param_properties['value/range'],
                                        param_properties['direction'],
                                        param_properties['description']])
        else:
                form.parameters = ['-','-','-','-','-',]
        return form

    def toExcel(self):
        self.fillHeader()
        ws = self.worksheet
        table_spacing = 3
        row = ws.max_row + table_spacing
        column = 2
        for section in self.xmlclass.sections:
            if section.kind in export_to_excel:
                for field in section.fields:
                    form = self.fillForm(field) 
                    table_start_row = row
                    ws.cell(row,1).value = form.header
                    ws.merge_cells(start_row=row, start_column=1, end_row=row,end_column=6)
                    ws.cell(row=row,column=1).alignment = Alignment(horizontal='center')
                    self.setAlignment(row,row,1,1, Alignment(horizontal='center',vertical='center',wrap_text=True))
                    row += 1
                    for k, v in form.table.items():
                        ws.cell(row, 1).value = k
                        ws.cell(row,1).alignment = Alignment(horizontal='center')
                        self.setAlignment(row,row,1,1, Alignment(horizontal='center',vertical='center',wrap_text=True))
                        if type(v) == list:
                            ws.merge_cells(start_row=row, start_column=1, end_row=row + len(v) - 1,end_column=1)
                            row_old = row
                            for inner in v:
                                for val in inner:
                                    if val in v[0]:
                                        ws.cell(row, column).fill = self.fill
                                        self.setAlignment(row,row,column,column, Alignment(horizontal='center',vertical='center',wrap_text=True))
                                        self.setFill(row,row,1,6,self.fill)
                                    else:
                                        self.setAlignment(row,row,column,column,Alignment(wrap_text=True))
                                    ws.cell(row, column).value = val
                                    column += 1
                                    if k == 'Return Type' or k == 'Global Variables':
                                        column +=1
                                column = 2
                                row+=1
                            if k == 'Return Type' or k == 'Global Variables':
                                merge_row = row_old
                                while merge_row < row:
                                    ws.merge_cells(start_row=merge_row, start_column=2, end_row=merge_row,end_column = 3)
                                    ws.merge_cells(start_row=merge_row, start_column=4, end_row=merge_row,end_column = 6)
                                    merge_row += 1
                        else:
                            ws.cell(row, column).value = v
                            self.setAlignment(row,row,column,column,Alignment(wrap_text=True))
                            ws.merge_cells(start_row=row, start_column=2, end_row=row,end_column = 6)
                            row +=1

                    self.setBorder(row_start = table_start_row, row_end=row-1,col_start=1,col_end=6,border=self.border)
                    self.setFill(table_start_row,row - 1,1,1,self.fill)
                    row += table_spacing

    def setBorder(self, row_start, row_end, col_start, col_end,border : Border):
        ws = self.worksheet
        for row in ws.iter_rows(min_row=row_start, min_col=col_start, max_row=row_end, max_col=col_end):
            for cell in row:
                cell.border = border
    def setAlignment(self, row_start, row_end, col_start, col_end,alignment):
        ws = self.worksheet
        for row in ws.iter_rows(min_row=row_start, min_col=col_start, max_row=row_end, max_col=col_end):
            for cell in row:
                cell.alignment = alignment

    def setFill(self, row_start, row_end, col_start, col_end,fill : PatternFill):
        ws = self.worksheet
        for row in ws.iter_rows(min_row=row_start, min_col=col_start, max_row=row_end, max_col=col_end):
            for cell in row:
                cell.fill = fill

def main():
    parser = argparse.ArgumentParser(
                    prog = 'ProgramName',
                    description = 'What the program does',
                    epilog = 'Text at the bottom of help')
    parser.add_argument("-d", "--Directory",dest="directory", help = "Starting directory",required=True)
    args = parser.parse_args()
    files = [file for file in glob.glob(args.directory + '/*.xml') if os.path.basename(file).startswith('class') and not os.path.basename(file).__contains__('Base')]
    wb = Workbook()
    wb.remove(wb.active)
    for file in files:
        xmlclass = XmlClass(file)
        tbl = Table(wb,xmlclass)
        tbl.toExcel()
    wb.save(os.path.join(os.getcwd(),'result.xlsx'))

if __name__ == '__main__':
    main()


